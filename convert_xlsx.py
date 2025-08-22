#!/usr/bin/env python3
import pandas as pd
import sys

try:
    # Leer el archivo Excel
    df = pd.read_excel("/workspaces/siigo-retiros/Untitled spreadsheet.xlsx")
    
    # Mostrar información básica
    print(f"Archivo cargado con {len(df)} filas y {len(df.columns)} columnas")
    print("\nPrimeras columnas:")
    print(list(df.columns)[:10])
    
    print("\nPrimeras 5 filas:")
    print(df.head())
    
    # Guardar como CSV
    df.to_csv("/workspaces/siigo-retiros/recuperacion_users.csv", index=False)
    print("\nArchivo convertido a: recuperacion_users.csv")
    
except ImportError:
    print("Pandas no está disponible. Intentando otra forma...")
    import openpyxl
    
    wb = openpyxl.load_workbook("/workspaces/siigo-retiros/Untitled spreadsheet.xlsx")
    ws = wb.active
    
    print(f"Archivo Excel cargado. Hoja activa: {ws.title}")
    print(f"Dimensiones: {ws.max_row} filas, {ws.max_column} columnas")
    
    # Mostrar primeras filas
    print("\nPrimeras 5 filas:")
    for row_num in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(11, ws.max_column + 1)):  # Primeras 10 columnas
            cell_value = ws.cell(row=row_num, column=col).value
            row_data.append(str(cell_value) if cell_value is not None else "")
        print(f"Fila {row_num}: {' | '.join(row_data)}")
    
except Exception as e:
    print(f"Error: {e}")
    print("Instalando pandas...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    
    # Reintentar
    import pandas as pd
    df = pd.read_excel("/workspaces/siigo-retiros/Untitled spreadsheet.xlsx")
    print(f"Archivo cargado con {len(df)} filas y {len(df.columns)} columnas")
    print(df.head())
    df.to_csv("/workspaces/siigo-retiros/recuperacion_users.csv", index=False)